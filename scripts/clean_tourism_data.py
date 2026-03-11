from __future__ import annotations

from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook


INPUT_XLSX = Path(r"C:\Users\12697\Downloads\Name your insight (2).xlsx")
WORKDIR = Path(__file__).resolve().parent
OUTPUT_XLSX = WORKDIR / "tourism_cleaned_outputs.xlsx"
OUTPUT_MONTHLY_CSV = WORKDIR / "tourism_monthly_clean.csv"
OUTPUT_TREE_CSV = WORKDIR / "tourism_decision_tree_ready.csv"
OUTPUT_SUMMARY_TXT = WORKDIR / "tourism_cleaning_summary.txt"


def load_source_dataframe(path: Path) -> pd.DataFrame:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))

    columns = [
        "date",
        "avg_stay_monthly",
        "hotel_occ",
        "spend_per_capita",
        "tourism_receipts",
        "avg_stay_annual",
        "visitor_arrivals",
        "visitor_arrivals_china",
    ]

    df = pd.DataFrame(rows[29:], columns=columns)
    df["date"] = pd.to_datetime(df["date"], errors="coerce")
    for col in columns[1:]:
        df[col] = pd.to_numeric(df[col], errors="coerce")

    return df.sort_values("date").reset_index(drop=True)


def assign_period(ts: pd.Timestamp) -> str | None:
    if pd.isna(ts):
        return None
    if ts <= pd.Timestamp("2020-01-01"):
        return "pre_covid"
    if ts <= pd.Timestamp("2021-12-01"):
        return "covid_shock"
    return "recovery"


def build_monthly_clean(df: pd.DataFrame) -> tuple[pd.DataFrame, float]:
    monthly = df.dropna(
        subset=[
            "date",
            "avg_stay_monthly",
            "hotel_occ",
            "visitor_arrivals",
            "visitor_arrivals_china",
        ]
    ).copy()

    monthly["year"] = monthly["date"].dt.year
    monthly["month"] = monthly["date"].dt.month
    monthly["quarter"] = monthly["date"].dt.quarter
    monthly["period"] = monthly["date"].apply(assign_period)
    monthly["china_share"] = monthly["visitor_arrivals_china"] / monthly["visitor_arrivals"]

    # Cap only the high-end pandemic spike to keep a tree from splitting on one extreme value.
    stay_cap = float(monthly["avg_stay_monthly"].quantile(0.95))
    monthly["avg_stay_monthly_capped"] = monthly["avg_stay_monthly"].clip(upper=stay_cap)

    monthly["visitor_arrivals_millions"] = monthly["visitor_arrivals"] / 1_000_000
    monthly["visitor_arrivals_china_thousands"] = monthly["visitor_arrivals_china"] / 1_000
    monthly["china_share_pct"] = monthly["china_share"] * 100

    return monthly, stay_cap


def build_tree_ready(monthly: pd.DataFrame) -> tuple[pd.DataFrame, float, float]:
    tree = monthly[
        [
            "date",
            "year",
            "month",
            "quarter",
            "period",
            "visitor_arrivals",
            "visitor_arrivals_china",
            "china_share",
            "china_share_pct",
            "hotel_occ",
            "avg_stay_monthly",
            "avg_stay_monthly_capped",
        ]
    ].copy()

    low_cut, high_cut = tree["hotel_occ"].quantile([1 / 3, 2 / 3]).tolist()

    tree["hotel_occ_level_tertile"] = pd.qcut(
        tree["hotel_occ"],
        q=3,
        labels=["low", "medium", "high"],
        duplicates="drop",
    ).astype(str)

    tree["hotel_occ_level_business"] = np.select(
        [
            tree["hotel_occ"] < 70,
            tree["hotel_occ"] <= 85,
            tree["hotel_occ"] > 85,
        ],
        [
            "low",
            "medium",
            "high",
        ],
        default="medium",
    )

    split_idx = int(np.floor(len(tree) * 0.8))
    tree["dataset_split"] = "train"
    tree.loc[tree.index >= split_idx, "dataset_split"] = "test"

    return tree, float(low_cut), float(high_cut)


def write_summary(
    source_df: pd.DataFrame,
    monthly: pd.DataFrame,
    tree: pd.DataFrame,
    stay_cap: float,
    low_cut: float,
    high_cut: float,
) -> None:
    lines = [
        "Tourism data cleaning summary",
        f"Source file: {INPUT_XLSX}",
        f"Raw rows in data block: {len(source_df)}",
        f"Monthly clean rows: {len(monthly)}",
        f"Tree-ready rows: {len(tree)}",
        f"Monthly date range: {monthly['date'].min().date()} to {monthly['date'].max().date()}",
        f"Avg stay cap (95th percentile): {stay_cap:.6f}",
        f"Hotel occupancy tertile low cut: {low_cut:.6f}",
        f"Hotel occupancy tertile high cut: {high_cut:.6f}",
        "",
        "Tree-ready columns:",
        ", ".join(tree.columns.tolist()),
    ]
    OUTPUT_SUMMARY_TXT.write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    source_df = load_source_dataframe(INPUT_XLSX)
    monthly, stay_cap = build_monthly_clean(source_df)
    tree, low_cut, high_cut = build_tree_ready(monthly)

    monthly.to_csv(OUTPUT_MONTHLY_CSV, index=False, encoding="utf-8-sig")
    tree.to_csv(OUTPUT_TREE_CSV, index=False, encoding="utf-8-sig")

    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
        monthly.to_excel(writer, sheet_name="monthly_clean", index=False)
        tree.to_excel(writer, sheet_name="decision_tree_ready", index=False)

    write_summary(source_df, monthly, tree, stay_cap, low_cut, high_cut)

    print(f"Created: {OUTPUT_MONTHLY_CSV}")
    print(f"Created: {OUTPUT_TREE_CSV}")
    print(f"Created: {OUTPUT_XLSX}")
    print(f"Created: {OUTPUT_SUMMARY_TXT}")


if __name__ == "__main__":
    main()
